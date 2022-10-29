import openpyxl
import logging


def is_intvalue_in_range(value_, range_):
    if isinstance(value_, int) and value_ in range_:
        return True
    else:
        return False


def is_positive_integer(value_):
    if isinstance(value_, int) and value_ >= 0:
        return True
    else:
        return False


def format_data_in_tuple(tuple_):
    list_ = [
        tuple_item.strip() if isinstance(tuple_item, str) else tuple_item
        for tuple_item in tuple_
    ]
    return tuple(list_)


def validate_range(min_, max_, range_upper_limit):
    assert is_positive_integer(range_upper_limit)
    limits_of_range = range(0, range_upper_limit + 1)
    if not is_intvalue_in_range(min_, limits_of_range):
        logging.error("The min value exceeds limit")
        return
    if max_ is not None and not is_intvalue_in_range(
        max_, limits_of_range
    ):
        logging.error("The max value exceeds limit")
        return
    if max_ is None:
        max_ = range_upper_limit
    if min_ == max_:
        return
    return min_, max_


class Work_Book:
    def __init__(self, path):
        self.path = path

    def load(self):
        try:
            # Define variable to load the wookbook
            workbook = openpyxl.load_workbook(self.path)
        except FileNotFoundError as err:
            logging.error(f"Error: File not found. Details:{err}")
            return
        except OSError as err:
            logging.error(f"Error: Can not open file. Details:{err}")
            return
        return workbook


class WorkSheet(Work_Book):
    def __init__(self, path, worksheet_name=None):
        self.worksheet_name = worksheet_name
        super().__init__(path)

    def read_worksheet(self):
        workbook = self.load()
        if not workbook:
            return
        if self.worksheet_name:
            try:
                worksheet = workbook[self.worksheet_name]
            except KeyError as err:
                logging.error(f"KeyError:{err}")
                return
        else:
            worksheet = workbook.active
        return worksheet


class TableRanges(WorkSheet):
    def __init__(
        self, path, worksheet_name=None, row_col_ranges_dict: dict = {}
    ):
        self.row_col_ranges_dict = row_col_ranges_dict
        super().__init__(path, worksheet_name)
        self.worksheet = self.read_worksheet()

    def get_row_col_ranges(self):
        if not self.row_col_ranges_dict:
            self.row_col_ranges_dict["min_row"] = 0
            self.row_col_ranges_dict["max_row"] = None
            self.row_col_ranges_dict["min_column"] = 0
            self.row_col_ranges_dict["max_column"] = None

    def validate_row_col_ranges(self):
        self.get_row_col_ranges()
        row_ = validate_range(
            self.row_col_ranges_dict["min_row"],
            self.row_col_ranges_dict["max_row"],
            self.worksheet.max_row,
        )
        table_ranges_dict = {}
        if row_:
            table_ranges_dict["min_row"], table_ranges_dict["max_row"] = row_
        else:
            return
        col_ = validate_range(
            self.row_col_ranges_dict["min_column"],
            self.row_col_ranges_dict["max_column"],
            self.worksheet.max_column,
        )
        if col_:
            table_ranges_dict["min_col"], table_ranges_dict["max_col"] = col_
        else:
            return
        return table_ranges_dict


class Data(TableRanges):
    def __init__(
        self, path, worksheet_name=None, row_col_ranges_dict: dict = {}
    ):
        super().__init__(path, worksheet_name, row_col_ranges_dict)

    def get_data_from_worksheet(self):
        worksheet = self.read_worksheet()
        if not worksheet:
            return

        worksheet_ranges = self.validate_row_col_ranges()
        if not worksheet_ranges:
            return

        worksheet_iterator = worksheet.iter_rows(
            min_row=worksheet_ranges["min_row"],
            max_row=worksheet_ranges["max_row"],
            min_col=worksheet_ranges["min_col"],
            max_col=worksheet_ranges["max_col"],
            values_only=True,
        )

        table_head = next(worksheet_iterator)
        table_head = format_data_in_tuple(table_head)
        data_list_of_dicts = []
        for value in worksheet_iterator:
            value = format_data_in_tuple(value)
            new_dict = dict(zip(table_head, value))
            data_list_of_dicts.append(new_dict)

        return data_list_of_dicts


if __name__ == "__main__":
    pass
